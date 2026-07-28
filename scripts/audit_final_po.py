#!/usr/bin/env python3
"""
TX PR Auditor - validate Final PO against create-pr-cd ECC output.

The auditor treats generated ECC rows as the expected entitlement. EPMS and
PR Model workbooks are intentionally not inputs here; create-pr-cd owns that
generation logic.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, replace
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple


SKILL_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DU_REGISTRY = SKILL_ROOT / "config" / "du_registry.json"


FINAL_PO_SHEET_NAME = "条目明细"
FINAL_PO_FORMATS = (
    ("条目明细", 1),
    ("Sheet1", 2),
)
ECC_SHEET_NAME = "details"

FINAL_PO_FIELD_MAP = {
    "派工日期": "dispatch_date",
    "æ´¾å·¥æ—¥æœŸ": "dispatch_date",
    "派工单号": "dispatch_order_number",
    "æ´¾å·¥å•å·": "dispatch_order_number",
    "PO行号": "po_line_number",
    "POè¡Œå·": "po_line_number",
    "需求单号": "request_number",
    "éœ€æ±‚å•å·": "request_number",
    "项目名称": "project_name",
    "é¡¹ç›®åç§°": "project_name",
    "项目编码": "project_code",
    "é¡¹ç›®ç¼–ç ": "project_code",
    "业务大类": "business_domain",
    "能力大类": "business_domain",
    "ä¸šåŠ¡å¤§ç±»": "business_domain",
    "施工区域": "region",
    "æ–½å·¥åŒºåŸŸ": "region",
    "采购区域": "purchasing_area",
    "é‡‡è´­åŒºåŸŸ": "purchasing_area",
    "分包商": "submitted_subcontractor",
    "åˆ†åŒ…å•†": "submitted_subcontractor",
    "逻辑站点名称": "logical_site_name",
    "é€»è¾‘ç«™ç‚¹åç§°": "logical_site_name",
    "逻辑站点编码": "du",
    "é€»è¾‘ç«™ç‚¹ç¼–ç ": "du",
    "物理站点名称": "physical_site_name",
    "ç‰©ç†ç«™ç‚¹åç§°": "physical_site_name",
    "物理站点编码": "site_code",
    "ç‰©ç†ç«™ç‚¹ç¼–ç ": "site_code",
    "外包代码": "submitted_item_code",
    "å¤–åŒ…ä»£ç ": "submitted_item_code",
    "代码名称": "submitted_item_description",
    "ä»£ç åç§°": "submitted_item_description",
    "量纲": "submitted_unit",
    "é‡çº²": "submitted_unit",
    "派工数量": "submitted_quantity",
    "æ´¾å·¥æ•°é‡": "submitted_quantity",
    "结算数量": "settlement_quantity",
    "ç»“ç®—æ•°é‡": "settlement_quantity",
    "支付数量": "paid_quantity",
    "æ”¯ä»˜æ•°é‡": "paid_quantity",
    "产品型号_备注": "product_model_remark",
    "äº§å“åž‹å·_å¤‡æ³¨": "product_model_remark",
    "派工单状态": "dispatch_status",
    "æ´¾å·¥å•çŠ¶æ€": "dispatch_status",
    "外包商编码": "subcontractor_code",
    "å¤–åŒ…å•†ç¼–ç ": "subcontractor_code",
}

ECC_FIELD_MAP = {
    "SN.": "sn",
    "Purchasing Area*": "purchasing_area",
    "Region*": "region",
    "Site ID*": "site_code",
    "Site Name*": "site_name",
    "Delivery Unit Code*": "du",
    "Logical Site Name": "logical_site_name",
    "Contract Number *": "contract_number",
    "Subcontractor*": "expected_subcontractor",
    "PBOM Code*": "expected_item_code",
    "SOW*": "expected_item_description",
    "Unit*": "expected_unit",
    "Quantity*": "expected_quantity",
    "Remarks": "remarks",
    "DU Model": "du_model_name",
    "DU Model Name": "du_model_name",
    "DU Model ID": "du_model_id",
    "DU Profile ID": "du_profile_id",
}

AUDIT_HEADERS = [
    "Source Row",
    "Scope",
    "DU Model",
    "DU Model ID",
    "DU Identity Key",
    "DU Profile ID",
    "Audit Result",
    "Reason Code",
    "Expected Item",
    "Expected Quantity",
    "Expected Subcontractor",
    "Normal Quantity",
    "Duplicate Quantity",
    "Expected ECC Evidence",
    "Matched ECC Evidence",
    "Explanation",
]

ECC_ANNOTATION_HEADERS = [
    "Audit Status",
    "Audit Reason Codes",
    "Final PO Match Count",
    "Submitted Quantity",
    "Normal Quantity",
    "Duplicate Quantity",
    "Final PO Evidence",
    "Audit Explanation",
]


@dataclass(frozen=True)
class FinalPORecord:
    source_row: int
    raw: Dict[str, Any]
    canonical: Dict[str, Any]


@dataclass(frozen=True)
class ExpectedECCRecord:
    source_file: str
    source_sheet: str
    source_row: int
    raw: Dict[str, Any]
    canonical: Dict[str, Any]


@dataclass(frozen=True)
class RawDataset:
    final_po_rows: List[Dict[str, Any]]
    ecc_rows: List[Dict[str, Any]]
    metadata: Dict[str, Any]


@dataclass(frozen=True)
class CanonicalDataset:
    final_po_records: List[FinalPORecord]
    expected_records: List[ExpectedECCRecord]
    metadata: Dict[str, Any]


@dataclass(frozen=True)
class ExpectedMatch:
    final_po: FinalPORecord
    scope: str
    du_model_name: str
    du_model_id: str
    du_identity_key: str
    du_profile_ids: str
    final_du_resolution_status: str
    expected_du_resolution_status: str
    candidate_site_key: str
    expected_items: List[ExpectedECCRecord]
    exact_item_candidates: List[ExpectedECCRecord]
    exact_item_records: List[ExpectedECCRecord]
    expected_subcontractors: List[str]
    evidence: str


@dataclass(frozen=True)
class AuditResult:
    final_po: FinalPORecord
    scope: str
    du_model_name: str
    du_model_id: str
    du_identity_key: str
    du_profile_ids: str
    classification: str
    reason_code: str
    expected_items: List[ExpectedECCRecord]
    expected_subcontractor: str
    expected_quantity: float
    normal_quantity: float
    duplicate_quantity: float
    explanation: str
    expected_evidence: str
    matched_evidence: str
    consumes_quantity: bool


@dataclass(frozen=True)
class AuditDataset:
    results: List[AuditResult]
    metadata: Dict[str, Any]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate Final PO.xlsx against create-pr-cd generated ECC output."
    )
    parser.add_argument("--final-po", default="input/Final PO.xlsx", help="Path to Final PO.xlsx")
    parser.add_argument(
        "--final-po-sheet",
        help="Final PO worksheet name. By default, auto-detect 条目明细 or Sheet1.",
    )
    parser.add_argument(
        "--final-po-header-row",
        type=int,
        help="1-based Final PO header row. By default, use row 1 for 条目明细 and row 2 for Sheet1.",
    )
    parser.add_argument("--final-po-max-rows", type=int, help="Optional maximum number of Final PO data rows to read")
    parser.add_argument("--filter-year", type=int, help="Audit only Final PO rows whose Dispatch Date matches this year")
    parser.add_argument("--filter-month", type=int, choices=range(1, 13), help="Audit only Final PO rows whose Dispatch Date matches this month")
    parser.add_argument(
        "--expected-ecc",
        action="append",
        required=True,
        help="Generated ECC .xlsx file or directory. Repeat for multiple paths.",
    )
    parser.add_argument(
        "--du-registry",
        default=str(DEFAULT_DU_REGISTRY),
        help="Nine-DU identity registry used to classify and isolate ECC entitlement.",
    )
    parser.add_argument("--ecc-sheet", default=ECC_SHEET_NAME, help="ECC worksheet name")
    parser.add_argument("--output", default="output/PR_Audit_Result.xlsx", help="Audit workbook path")
    parser.add_argument("--summary-json", help="Optional JSON summary path")
    parser.add_argument(
        "--annotate-ecc-output",
        action="store_true",
        help="Copy generated ECC workbooks to a timestamped folder and append audit status columns.",
    )
    parser.add_argument(
        "--annotated-ecc-output-root",
        default="output",
        help="Root directory for timestamped annotated ECC output folders.",
    )
    parser.add_argument(
        "--annotated-ecc-timestamp",
        help="Optional timestamp folder name for annotated ECC output, useful for deterministic tests.",
    )
    return parser.parse_args()


def require_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill
    except ModuleNotFoundError as exc:
        raise SystemExit("Missing dependency: openpyxl. Install with `python -m pip install -r requirements.txt`.") from exc
    return Workbook, load_workbook, Font, PatternFill


def require_file(path: Path, label: str) -> Path:
    if not path.exists():
        raise FileNotFoundError(f"{label} not found: {path}")
    if not path.is_file():
        raise FileNotFoundError(f"{label} is not a file: {path}")
    return path


def load_du_registry(path: Path = DEFAULT_DU_REGISTRY) -> Dict[str, Any]:
    registry_path = require_file(Path(path), "DU registry")
    registry = json.loads(registry_path.read_text(encoding="utf-8"))
    identities = registry.get("identities")
    if not isinstance(identities, list) or not identities:
        raise ValueError(f"DU registry has no identities: {registry_path}")

    identity_keys: set[str] = set()
    model_ids: set[str] = set()
    for identity in identities:
        if not isinstance(identity, dict):
            raise ValueError(f"DU registry identity must be an object: {registry_path}")
        required = ("identity_key", "project_key", "du_model_name", "du_model_id", "aliases", "profile_ids")
        missing = [field for field in required if not identity.get(field)]
        if missing:
            raise ValueError(f"DU registry identity is missing {', '.join(missing)}: {registry_path}")
        identity_key = text(identity["identity_key"])
        model_id = text(identity["du_model_id"])
        if identity_key in identity_keys:
            raise ValueError(f"Duplicate DU identity key in registry: {identity_key}")
        if model_id in model_ids:
            raise ValueError(f"Duplicate DU model ID in registry: {model_id}")
        identity_keys.add(identity_key)
        model_ids.add(model_id)

    return {
        **registry,
        "path": str(registry_path),
        "identities": identities,
    }


def normalize_du_phrase(value: Any) -> str:
    return " ".join(re.findall(r"[A-Z0-9]+", text(value).upper()))


def resolve_du_identity(
    registry: Dict[str, Any],
    *,
    du_model_name: Any = "",
    du_model_id: Any = "",
    du_profile_id: Any = "",
    source_text: Any = "",
) -> Optional[Dict[str, Any]]:
    identity, status = resolve_du_identity_evidence(
        registry,
        du_model_name=du_model_name,
        du_model_id=du_model_id,
        du_profile_id=du_profile_id,
        source_text=source_text,
    )
    return identity if status == "RESOLVED" else None


def resolve_du_identity_evidence(
    registry: Dict[str, Any],
    *,
    du_model_name: Any = "",
    du_model_id: Any = "",
    du_profile_id: Any = "",
    source_text: Any = "",
) -> Tuple[Optional[Dict[str, Any]], str]:
    identities = registry["identities"]
    matches: Dict[str, Dict[str, Any]] = {}
    unknown_explicit_evidence = False

    model_id = text(du_model_id)
    if model_id:
        identity = next(
            (item for item in identities if text(item["du_model_id"]) == model_id),
            None,
        )
        if identity is None:
            unknown_explicit_evidence = True
        else:
            matches[text(identity["identity_key"])] = identity

    profile_id = text(du_profile_id)
    if profile_id:
        identity = next(
            (
                item
                for item in identities
                if profile_id in {text(value) for value in item.get("profile_ids", [])}
            ),
            None,
        )
        if identity is None:
            unknown_explicit_evidence = True
        else:
            matches[text(identity["identity_key"])] = identity

    for value, explicit in ((du_model_name, True), (source_text, False)):
        value_text = text(value)
        if not value_text:
            continue
        haystack = normalize_du_phrase(value_text)
        value_matches: Dict[str, Dict[str, Any]] = {}
        for identity in identities:
            aliases = list(identity.get("aliases", [])) + [identity.get("du_model_name", "")]
            if any(
                (needle := normalize_du_phrase(alias))
                and f" {needle} " in f" {haystack} "
                for alias in aliases
            ):
                value_matches[text(identity["identity_key"])] = identity
        if explicit and not value_matches:
            unknown_explicit_evidence = True
        matches.update(value_matches)

    if len(matches) > 1 or (matches and unknown_explicit_evidence):
        return None, "CONFLICT"
    if not matches:
        return None, "UNKNOWN"
    return next(iter(matches.values())), "RESOLVED"


def apply_du_identity(
    data: Dict[str, Any],
    identity: Optional[Dict[str, Any]],
    resolution_status: Optional[str] = None,
) -> Dict[str, Any]:
    data["du_resolution_status"] = resolution_status or ("RESOLVED" if identity else "UNKNOWN")
    if identity is None:
        data["du_identity_key"] = ""
        data["du_model_name"] = text(data.get("du_model_name"))
        data["du_model_id"] = text(data.get("du_model_id"))
        data["du_profile_ids"] = text(data.get("du_profile_id"))
        return data
    data["du_identity_key"] = text(identity["identity_key"])
    data["du_model_name"] = text(identity["du_model_name"])
    data["du_model_id"] = text(identity["du_model_id"])
    data["du_profile_ids"] = join_unique(identity.get("profile_ids", []))
    return data


def expand_ecc_paths(paths: Sequence[str]) -> List[Path]:
    files: List[Path] = []
    for raw_path in paths:
        path = Path(raw_path)
        if not path.exists():
            raise FileNotFoundError(f"Expected ECC path not found: {path}")
        if path.is_dir():
            files.extend(sorted(p for p in path.iterdir() if p.suffix.lower() in {".xlsx", ".xlsm"} and not p.name.startswith("~$")))
        elif path.suffix.lower() in {".xlsx", ".xlsm"}:
            files.append(path)
        else:
            raise ValueError(f"Expected ECC input must be an .xlsx/.xlsm file or directory: {path}")
    unique: Dict[str, Path] = {}
    for file_path in files:
        unique[str(file_path.resolve())] = file_path
    if not unique:
        raise FileNotFoundError("No generated ECC .xlsx/.xlsm files were found.")
    return list(unique.values())


def normalize_header(value: Any) -> str:
    return str(value or "").strip()


def unique_headers(headers: Sequence[Any]) -> List[str]:
    counts: Dict[str, int] = defaultdict(int)
    out: List[str] = []
    for header in headers:
        name = normalize_header(header) or "EMPTY"
        counts[name] += 1
        out.append(name if counts[name] == 1 else f"{name}__{counts[name]}")
    return out


def trim_trailing_empty(values: Sequence[Any]) -> List[Any]:
    last_nonempty = 0
    for idx, value in enumerate(values, 1):
        if value not in (None, ""):
            last_nonempty = idx
    return list(values[:last_nonempty])


def read_table(
    path: Path,
    sheet_name: str,
    header_row: int,
    max_data_rows: Optional[int] = None,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    _, load_workbook, _, _ = require_openpyxl()
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Required worksheet '{sheet_name}' not found in {path}. "
            f"Available sheets: {', '.join(wb.sheetnames)}"
        )
    ws = wb[sheet_name]
    header_values = trim_trailing_empty(
        [cell.value for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
    )
    headers = unique_headers(header_values)
    rows: List[Dict[str, Any]] = []
    consecutive_blank_rows = 0
    max_blank_tail_rows = 1000
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        if max_data_rows is not None and len(rows) >= max_data_rows:
            break
        values = list(row[: len(headers)])
        if not any(value not in (None, "") for value in values):
            consecutive_blank_rows += 1
            if consecutive_blank_rows >= max_blank_tail_rows:
                break
            continue
        consecutive_blank_rows = 0
        record = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        record["_source_row"] = row_idx
        record["_source_file"] = str(path)
        record["_source_sheet"] = ws.title
        rows.append(record)
    metadata = {
        "path": str(path),
        "sheet": ws.title,
        "header_row": header_row,
        "row_count": len(rows),
        "max_data_rows": max_data_rows,
        "column_count": len(headers),
    }
    wb.close()
    return rows, metadata


def resolve_final_po_layout(
    path: Path,
    requested_sheet: Optional[str],
    requested_header_row: Optional[int],
) -> Tuple[str, int]:
    _, load_workbook, _, _ = require_openpyxl()
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if requested_sheet:
            if requested_sheet not in wb.sheetnames:
                raise ValueError(
                    f"Required worksheet '{requested_sheet}' not found in {path}. "
                    f"Available sheets: {', '.join(wb.sheetnames)}"
                )
            sheet_name = requested_sheet
        else:
            sheet_name = next(
                (candidate for candidate, _ in FINAL_PO_FORMATS if candidate in wb.sheetnames),
                "",
            )
            if not sheet_name:
                supported = ", ".join(name for name, _ in FINAL_PO_FORMATS)
                raise ValueError(
                    f"No supported Final PO worksheet found in {path}. "
                    f"Expected one of: {supported}. Available sheets: {', '.join(wb.sheetnames)}"
                )

        header_row = requested_header_row or dict(FINAL_PO_FORMATS).get(sheet_name, 1)
        if header_row < 1:
            raise ValueError("Final PO header row must be 1 or greater.")
        return sheet_name, header_row
    finally:
        wb.close()


def workbook_reader(
    final_po: Path,
    expected_ecc_files: Sequence[Path],
    final_po_sheet: Optional[str],
    final_po_header_row: Optional[int],
    final_po_max_rows: Optional[int],
    ecc_sheet: str,
) -> RawDataset:
    final_po_sheet, final_po_header_row = resolve_final_po_layout(
        final_po,
        final_po_sheet,
        final_po_header_row,
    )
    final_po_rows, final_meta = read_table(final_po, final_po_sheet, final_po_header_row, final_po_max_rows)
    ecc_rows: List[Dict[str, Any]] = []
    ecc_meta = []
    for ecc_file in expected_ecc_files:
        rows, meta = read_table(ecc_file, ecc_sheet, 1)
        ecc_rows.extend(rows)
        ecc_meta.append(meta)
    return RawDataset(
        final_po_rows=final_po_rows,
        ecc_rows=ecc_rows,
        metadata={
            "final_po": final_meta,
            "expected_ecc": ecc_meta,
            "expected_ecc_files": [str(path) for path in expected_ecc_files],
        },
    )


def canonicalize(raw: Dict[str, Any], field_map: Dict[str, str]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for source_field, canonical_field in field_map.items():
        if canonical_field not in out or out[canonical_field] in (None, ""):
            out[canonical_field] = raw.get(source_field)
    return out


def field_mapper(raw: RawDataset) -> CanonicalDataset:
    final_records = [
        FinalPORecord(
            source_row=int(row["_source_row"]),
            raw=row,
            canonical=canonicalize(row, FINAL_PO_FIELD_MAP),
        )
        for row in raw.final_po_rows
    ]
    expected_records = [
        ExpectedECCRecord(
            source_file=text(row["_source_file"]),
            source_sheet=text(row["_source_sheet"]),
            source_row=int(row["_source_row"]),
            raw=row,
            canonical=canonicalize(row, ECC_FIELD_MAP),
        )
        for row in raw.ecc_rows
    ]
    return CanonicalDataset(final_records, expected_records, raw.metadata)


def canonical_builder(
    dataset: CanonicalDataset,
    du_registry: Optional[Dict[str, Any]] = None,
) -> CanonicalDataset:
    du_registry = du_registry or load_du_registry()
    final_records: List[FinalPORecord] = []
    for record in dataset.final_po_records:
        data = dict(record.canonical)
        data["site_code"] = normalize_code(data.get("site_code"))
        data["du"] = normalize_code(data.get("du"))
        data["submitted_item_code"] = normalize_code(data.get("submitted_item_code"))
        data["submitted_quantity"] = submitted_quantity(data)
        data["submitted_subcontractor_norm"] = normalize_subcontractor(data.get("submitted_subcontractor"))
        data["dispatch_sort_key"] = dispatch_sort_key(data)
        final_identity, final_resolution_status = resolve_du_identity_evidence(
            du_registry,
            source_text=" ".join(
                text(data.get(field))
                for field in ("project_name", "product_model_remark")
                if text(data.get(field))
            ),
        )
        apply_du_identity(data, final_identity, final_resolution_status)
        final_records.append(replace(record, canonical=data))

    expected_records: List[ExpectedECCRecord] = []
    for record in dataset.expected_records:
        data = dict(record.canonical)
        data["site_code"] = normalize_code(data.get("site_code"))
        data["du"] = normalize_code(data.get("du"))
        data["expected_item_code"] = normalize_code(data.get("expected_item_code"))
        data["expected_quantity"] = to_float(data.get("expected_quantity"), default=0.0)
        data["expected_subcontractor_norm"] = normalize_subcontractor(data.get("expected_subcontractor"))
        data["scope"] = infer_scope_from_ecc(record.source_file)
        expected_identity, expected_resolution_status = resolve_du_identity_evidence(
            du_registry,
            du_model_name=data.get("du_model_name"),
            du_model_id=data.get("du_model_id"),
            du_profile_id=data.get("du_profile_id"),
            source_text=Path(record.source_file).name,
        )
        apply_du_identity(data, expected_identity, expected_resolution_status)
        data["entitlement_key"] = entitlement_key(data)
        expected_records.append(replace(record, canonical=data))

    metadata = dict(dataset.metadata)
    file_identities: Dict[str, Dict[str, str]] = {}
    for record in expected_records:
        file_name = str(Path(record.source_file))
        if file_name not in file_identities:
            file_identities[file_name] = {
                "du_model_name": text(record.canonical.get("du_model_name")) or "UNKNOWN",
                "du_model_id": text(record.canonical.get("du_model_id")),
                "du_identity_key": text(record.canonical.get("du_identity_key")),
                "du_profile_ids": text(record.canonical.get("du_profile_ids")),
            }
    identified_files = sum(bool(item["du_identity_key"]) for item in file_identities.values())
    metadata["du_support"] = {
        "registry_path": du_registry["path"],
        "supported_du_count": len(du_registry["identities"]),
        "identified_ecc_file_count": identified_files,
        "unknown_ecc_file_count": len(file_identities) - identified_files,
        "ecc_files": file_identities,
    }
    return CanonicalDataset(final_records, expected_records, metadata)


def filter_final_po_period(dataset: CanonicalDataset, year: Optional[int], month: Optional[int]) -> CanonicalDataset:
    if year is None and month is None:
        return dataset
    if year is None or month is None:
        raise ValueError("Final PO year and month filters must be provided together.")

    filtered_records = []
    for record in dataset.final_po_records:
        parsed = sortable_date(record.canonical.get("dispatch_date"))
        if isinstance(parsed, datetime) and parsed.year == year and parsed.month == month:
            filtered_records.append(record)

    metadata = dict(dataset.metadata)
    metadata["final_po_period_filter"] = {
        "year": year,
        "month": month,
        "input_row_count": len(dataset.final_po_records),
        "matched_row_count": len(filtered_records),
    }
    return CanonicalDataset(filtered_records, dataset.expected_records, metadata)


def expected_matcher(dataset: CanonicalDataset) -> List[ExpectedMatch]:
    expected_by_site: Dict[str, List[ExpectedECCRecord]] = defaultdict(list)
    for expected in dataset.expected_records:
        for key in site_keys(expected.canonical):
            expected_by_site[key].append(expected)

    matches: List[ExpectedMatch] = []
    for final_po in dataset.final_po_records:
        f = final_po.canonical
        candidate_keys = [key for key in [f.get("site_code"), f.get("du")] if key]
        expected_items = []
        site_key = ""
        final_du_identity = text(f.get("du_identity_key"))
        final_scope = infer_scope_from_final_po(final_po)
        for key in candidate_keys:
            candidates = expected_by_site.get(key, [])
            if final_du_identity:
                candidates = [
                    item
                    for item in candidates
                    if text(item.canonical.get("du_identity_key")) == final_du_identity
                ]
            if final_scope != "UNKNOWN":
                candidates = [
                    item
                    for item in candidates
                    if text(item.canonical.get("scope")) == final_scope
                ]
            if candidates:
                expected_items = candidates
                site_key = key
                break
        submitted_code = f.get("submitted_item_code")
        exact_candidates = [
            item
            for item in expected_items
            if item.canonical.get("expected_item_code") == submitted_code
        ]
        submitted_subcon = text(f.get("submitted_subcontractor_norm"))
        exact = [
            item
            for item in exact_candidates
            if not submitted_subcon
            or text(item.canonical.get("expected_subcontractor_norm")) == submitted_subcon
        ]
        scope = infer_scope_from_expected_or_final(exact_candidates or expected_items, final_po)
        identity_records = exact_candidates or expected_items
        identity_keys = {
            text(item.canonical.get("du_identity_key")) or "UNKNOWN"
            for item in identity_records
        }
        if len(identity_keys) > 1:
            du_identity_key = "MULTI"
            du_model_name = "MULTI"
            du_model_id = "MULTI"
        elif identity_records:
            du_identity_key = next(iter(identity_keys))
            du_model_name = join_unique(item.canonical.get("du_model_name") for item in identity_records) or "UNKNOWN"
            du_model_id = join_unique(item.canonical.get("du_model_id") for item in identity_records)
        else:
            du_identity_key = final_du_identity
            du_model_name = text(f.get("du_model_name"))
            du_model_id = text(f.get("du_model_id"))
        du_profile_ids = join_unique(
            item.canonical.get("du_profile_ids")
            for item in identity_records
        )
        subcon = sorted({text(item.canonical.get("expected_subcontractor")) for item in expected_items if text(item.canonical.get("expected_subcontractor"))})
        matches.append(
            ExpectedMatch(
                final_po=final_po,
                scope=scope,
                du_model_name=du_model_name,
                du_model_id=du_model_id,
                du_identity_key=du_identity_key,
                du_profile_ids=du_profile_ids,
                final_du_resolution_status=text(f.get("du_resolution_status")) or "UNKNOWN",
                expected_du_resolution_status=expected_du_resolution_status(identity_records),
                candidate_site_key=site_key,
                expected_items=expected_items,
                exact_item_candidates=exact_candidates,
                exact_item_records=exact,
                expected_subcontractors=subcon,
                evidence=expected_evidence(expected_items),
            )
        )
    return matches


def audit_engine(matches: Sequence[ExpectedMatch], metadata: Dict[str, Any]) -> AuditDataset:
    results: List[AuditResult] = []
    for match in matches:
        final_po = match.final_po
        f = final_po.canonical
        submitted_subcon = f.get("submitted_subcontractor_norm")
        exact_subcons = {
            item.canonical.get("expected_subcontractor_norm")
            for item in match.exact_item_candidates
            if item.canonical.get("expected_subcontractor_norm")
        }
        site_subcons = {
            item.canonical.get("expected_subcontractor_norm")
            for item in match.expected_items
            if item.canonical.get("expected_subcontractor_norm")
        }
        expected_subcontractor = join_unique(
            item.canonical.get("expected_subcontractor")
            for item in match.exact_item_records
        ) or "; ".join(match.expected_subcontractors)

        if not match.expected_items:
            results.append(
                make_result(
                    match,
                    "Abnormal - Invalid PO",
                    "INVALID_NOT_IN_CREATE_PR_CD_OUTPUT",
                    "No generated ECC entitlement exists for the submitted site or DU.",
                    expected_subcontractor,
                    consumes=False,
                )
            )
            continue

        if (
            match.final_du_resolution_status == "CONFLICT"
            or match.expected_du_resolution_status == "CONFLICT"
        ):
            results.append(
                make_result(
                    match,
                    "Abnormal - Invalid PO",
                    "INVALID_CONFLICTING_DU_IDENTITY",
                    "DU identity evidence conflicts across registered model, profile, or filename signals.",
                    expected_subcontractor,
                    consumes=False,
                )
            )
            continue

        if match.expected_du_resolution_status == "UNKNOWN":
            results.append(
                make_result(
                    match,
                    "Abnormal - Invalid PO",
                    "INVALID_UNKNOWN_DU_MODEL",
                    "Generated ECC entitlement is not associated with a registered create-pr-cd DU identity.",
                    expected_subcontractor,
                    consumes=False,
                )
            )
            continue

        if match.du_identity_key == "MULTI":
            results.append(
                make_result(
                    match,
                    "Abnormal - Invalid PO",
                    "INVALID_AMBIGUOUS_DU_MODEL",
                    "The submitted site and item match entitlement from multiple DU models; identify the DU model in Final PO.",
                    expected_subcontractor,
                    consumes=False,
                )
            )
            continue

        if match.scope == "MULTI":
            results.append(
                make_result(
                    match,
                    "Abnormal - Invalid PO",
                    "INVALID_AMBIGUOUS_SCOPE",
                    "The submitted site and item match entitlement from multiple scopes; identify the scope in Final PO.",
                    expected_subcontractor,
                    consumes=False,
                )
            )
            continue

        if submitted_subcon and site_subcons and submitted_subcon not in site_subcons:
            results.append(
                make_result(
                    match,
                    "Abnormal - Invalid PO",
                    "INVALID_SUBCON_CHANGED",
                    "Submitted subcontractor does not match the generated ECC subcontractor for this site.",
                    expected_subcontractor,
                    consumes=False,
                )
            )
            continue

        if not match.exact_item_candidates:
            results.append(
                make_result(
                    match,
                    "Abnormal - Wrong PO",
                    "WRONG_LINE_ITEM_MAPPING",
                    "Submitted item code is not present in the generated ECC entitlement for this site.",
                    expected_subcontractor,
                    consumes=False,
                )
            )
            continue

        if submitted_subcon and exact_subcons and submitted_subcon not in exact_subcons:
            results.append(
                make_result(
                    match,
                    "Abnormal - Invalid PO",
                    "INVALID_SUBCON_CHANGED",
                    "Submitted subcontractor does not match the generated ECC subcontractor for this item.",
                    expected_subcontractor,
                    consumes=False,
                )
            )
            continue

        if not submitted_subcon and len(exact_subcons) > 1:
            results.append(
                make_result(
                    match,
                    "Abnormal - Invalid PO",
                    "INVALID_AMBIGUOUS_SUBCONTRACTOR",
                    "The submitted item has entitlement from multiple subcontractors; identify the subcontractor in Final PO.",
                    expected_subcontractor,
                    consumes=False,
                )
            )
            continue

        expected_qty = sum(item.canonical.get("expected_quantity", 0.0) for item in match.exact_item_records)
        results.append(
            make_result(
                match,
                "PENDING_QUANTITY",
                "PENDING_QUANTITY",
                "Submitted item matches generated ECC entitlement and is ready for quantity consumption.",
                expected_subcontractor,
                expected_quantity=expected_qty,
                consumes=True,
            )
        )
    return AuditDataset(results, metadata)


def duplicate_resolver(dataset: AuditDataset) -> AuditDataset:
    results = list(dataset.results)
    pending = [idx for idx, result in enumerate(results) if result.consumes_quantity]
    pending.sort(key=lambda idx: results[idx].final_po.canonical.get("dispatch_sort_key"))

    consumed: Dict[Tuple[str, str, str, str, str], float] = defaultdict(float)
    for idx in pending:
        result = results[idx]
        f = result.final_po.canonical
        key = consumption_key(result)
        submitted_qty = f.get("submitted_quantity", 0.0)
        available = max(result.expected_quantity - consumed[key], 0.0)
        normal_qty = min(submitted_qty, available)
        duplicate_qty = max(submitted_qty - normal_qty, 0.0)
        consumed[key] += normal_qty
        if duplicate_qty <= 0:
            results[idx] = replace(
                result,
                classification="Normal",
                reason_code="NORMAL_FULL",
                normal_quantity=normal_qty,
                duplicate_quantity=0.0,
                explanation="Submitted quantity is within generated ECC entitlement.",
                consumes_quantity=False,
            )
        elif normal_qty > 0:
            results[idx] = replace(
                result,
                classification="Abnormal - Duplicate PO",
                reason_code="DUPLICATE_PARTIAL_QUANTITY",
                normal_quantity=normal_qty,
                duplicate_quantity=duplicate_qty,
                explanation="Part of the submitted quantity exceeds generated ECC entitlement already consumed in this snapshot.",
                consumes_quantity=False,
            )
        else:
            results[idx] = replace(
                result,
                classification="Abnormal - Duplicate PO",
                reason_code="DUPLICATE_FULL_QUANTITY",
                normal_quantity=0.0,
                duplicate_quantity=duplicate_qty,
                explanation="Submitted quantity exceeds generated ECC entitlement already consumed in this snapshot.",
                consumes_quantity=False,
            )
    return AuditDataset(results, dataset.metadata)


def report_writer(dataset: AuditDataset, output: Path, summary_json: Optional[Path]) -> Dict[str, Any]:
    Workbook, _, Font, PatternFill = require_openpyxl()
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "PR_Audit_Result"

    raw_headers = source_headers(dataset.results)
    headers = raw_headers + AUDIT_HEADERS
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.font = Font(bold=True)
        if col_idx > len(raw_headers):
            cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    for row_idx, result in enumerate(dataset.results, 2):
        for col_idx, header in enumerate(raw_headers, 1):
            ws.cell(row_idx, col_idx, result.final_po.raw.get(header))
        audit_values = [
            result.final_po.source_row,
            result.scope,
            result.du_model_name,
            result.du_model_id,
            result.du_identity_key,
            result.du_profile_ids,
            result.classification,
            result.reason_code,
            expected_item_text(result.expected_items),
            result.expected_quantity,
            result.expected_subcontractor,
            result.normal_quantity,
            result.duplicate_quantity,
            result.expected_evidence,
            result.matched_evidence,
            result.explanation,
        ]
        for offset, value in enumerate(audit_values, 1):
            ws.cell(row_idx, len(raw_headers) + offset, value)

    for idx, header in enumerate(headers, 1):
        width = min(max(len(str(header)) + 2, 12), 48)
        ws.column_dimensions[column_name(idx)].width = width
    wb.save(output)

    summary = {
        "total_rows": len(dataset.results),
        "classifications": dict(Counter(result.classification for result in dataset.results)),
        "reason_codes": dict(Counter(result.reason_code for result in dataset.results)),
        "du_models": dict(Counter(result.du_model_name or "UNKNOWN" for result in dataset.results)),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "metadata": dataset.metadata,
    }
    if summary_json:
        summary_json.parent.mkdir(parents=True, exist_ok=True)
        summary_json.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    return summary


def annotated_ecc_writer(
    dataset: AuditDataset,
    expected_ecc_files: Sequence[Path],
    output_root: Path,
    timestamp: Optional[str],
    ecc_sheet: str,
) -> Dict[str, Any]:
    _, load_workbook, Font, PatternFill = require_openpyxl()
    run_timestamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = output_root / run_timestamp
    output_dir.mkdir(parents=True, exist_ok=True)

    annotation_index = build_ecc_annotation_index(dataset.results)
    seen_output_names: set[str] = set()
    file_summaries: List[Dict[str, Any]] = []
    status_counts: Counter[str] = Counter()
    reason_counts: Counter[str] = Counter()

    for source_file in expected_ecc_files:
        output_name = source_file.name
        if output_name in seen_output_names:
            raise ValueError(f"Duplicate annotated ECC output filename would overwrite another file: {output_name}")
        seen_output_names.add(output_name)

        wb = load_workbook(source_file)
        if ecc_sheet not in wb.sheetnames:
            raise ValueError(
                f"Required worksheet '{ecc_sheet}' not found in {source_file}. "
                f"Available sheets: {', '.join(wb.sheetnames)}"
            )
        ws = wb[ecc_sheet]
        start_col = ws.max_column + 1
        header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for offset, header in enumerate(ECC_ANNOTATION_HEADERS):
            cell = ws.cell(1, start_col + offset, header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            width = min(max(len(header) + 2, 14), 42)
            ws.column_dimensions[column_name(start_col + offset)].width = width

        file_status_counts: Counter[str] = Counter()
        annotated_rows = 0
        for row_idx in range(2, ws.max_row + 1):
            key = ecc_annotation_key(source_file.name, ws.title, row_idx)
            values = annotation_values(annotation_index.get(key, []))
            for offset, value in enumerate(values):
                ws.cell(row_idx, start_col + offset, value)
            status = text(values[0])
            file_status_counts[status] += 1
            status_counts[status] += 1
            for reason in split_joined(values[1]):
                reason_counts[reason] += 1
            annotated_rows += 1

        output_file = output_dir / output_name
        wb.save(output_file)
        wb.close()
        du_file_metadata = (
            dataset.metadata.get("du_support", {})
            .get("ecc_files", {})
            .get(str(source_file), {})
        )
        file_summaries.append(
            {
                "source_file": str(source_file),
                "output_file": str(output_file),
                "worksheet": ecc_sheet,
                "du_model_name": du_file_metadata.get("du_model_name", "UNKNOWN"),
                "du_model_id": du_file_metadata.get("du_model_id", ""),
                "du_identity_key": du_file_metadata.get("du_identity_key", ""),
                "du_profile_ids": du_file_metadata.get("du_profile_ids", ""),
                "annotated_rows": annotated_rows,
                "status_counts": dict(file_status_counts),
            }
        )

    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "output_dir": str(output_dir),
        "file_count": len(file_summaries),
        "status_counts": dict(status_counts),
        "reason_codes": dict(reason_counts),
        "files": file_summaries,
    }
    summary_path = output_dir / "annotated_ecc.summary.json"
    summary_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    return summary


def build_ecc_annotation_index(results: Sequence[AuditResult]) -> Dict[Tuple[str, str, int], List[AuditResult]]:
    index: Dict[Tuple[str, str, int], List[AuditResult]] = defaultdict(list)
    for result in results:
        for key in parse_ecc_evidence_keys(result.matched_evidence):
            index[key].append(result)
    return index


def ecc_annotation_key(source_file_name: str, source_sheet: str, source_row: int) -> Tuple[str, str, int]:
    return (source_file_name, source_sheet, int(source_row))


def parse_ecc_evidence_keys(evidence: str) -> List[Tuple[str, str, int]]:
    keys: List[Tuple[str, str, int]] = []
    for part in split_joined(evidence):
        if part.startswith("... +"):
            continue
        match = re.match(r"(.+):([^:!]+)!(\d+)$", part)
        if match:
            keys.append(ecc_annotation_key(match.group(1), match.group(2), int(match.group(3))))
    return keys


def annotation_values(results: Sequence[AuditResult]) -> List[Any]:
    if not results:
        return ["NOT_IN_FINAL_PO", "", 0, 0.0, 0.0, 0.0, "", ""]

    return [
        aggregate_annotation_status(results),
        join_unique(result.reason_code for result in results),
        len(results),
        sum(result.final_po.canonical.get("submitted_quantity", 0.0) for result in results),
        sum(result.normal_quantity for result in results),
        sum(result.duplicate_quantity for result in results),
        join_unique(final_po_evidence(result) for result in results),
        join_unique(result.explanation for result in results),
    ]


def aggregate_annotation_status(results: Sequence[AuditResult]) -> str:
    statuses = {annotation_status(result.classification) for result in results}
    if len(statuses) > 1:
        return "MIXED"
    return next(iter(statuses))


def annotation_status(classification: str) -> str:
    if classification == "Normal":
        return "NORMAL"
    if classification == "Abnormal - Duplicate PO":
        return "DUPLICATE"
    if classification == "Abnormal - Invalid PO":
        return "INVALID"
    if classification == "Abnormal - Wrong PO":
        return "WRONG"
    return text(classification).upper().replace(" ", "_") or "UNKNOWN"


def final_po_evidence(result: AuditResult) -> str:
    data = result.final_po.canonical
    parts = [
        f"Source Row={result.final_po.source_row}",
        f"Request Number={text(data.get('request_number'))}",
        f"Dispatch Order Number={text(data.get('dispatch_order_number'))}",
        f"PO Line Number={text(data.get('po_line_number'))}",
    ]
    return "; ".join(part for part in parts if not part.endswith("="))


def join_unique(values: Iterable[Any]) -> str:
    seen: set[str] = set()
    out: List[str] = []
    for value in values:
        value_text = text(value)
        if value_text and value_text not in seen:
            seen.add(value_text)
            out.append(value_text)
    return "; ".join(out)


def split_joined(value: Any) -> List[str]:
    return [part.strip() for part in text(value).split(";") if part.strip()]


def run_pipeline(args: argparse.Namespace) -> Dict[str, Any]:
    final_po = require_file(Path(args.final_po), "Final PO")
    ecc_files = expand_ecc_paths(args.expected_ecc)
    du_registry = load_du_registry(Path(args.du_registry))
    raw = workbook_reader(
        final_po=final_po,
        expected_ecc_files=ecc_files,
        final_po_sheet=args.final_po_sheet,
        final_po_header_row=args.final_po_header_row,
        final_po_max_rows=args.final_po_max_rows,
        ecc_sheet=args.ecc_sheet,
    )
    mapped = field_mapper(raw)
    canonical = filter_final_po_period(
        canonical_builder(mapped, du_registry),
        args.filter_year,
        args.filter_month,
    )
    matches = expected_matcher(canonical)
    audited = audit_engine(matches, canonical.metadata)
    duplicated = duplicate_resolver(audited)
    summary = report_writer(
        duplicated,
        Path(args.output),
        Path(args.summary_json) if args.summary_json else None,
    )
    if args.annotate_ecc_output:
        summary["annotated_ecc"] = annotated_ecc_writer(
            duplicated,
            ecc_files,
            Path(args.annotated_ecc_output_root),
            args.annotated_ecc_timestamp,
            args.ecc_sheet,
        )
    return summary


def make_result(
    match: ExpectedMatch,
    classification: str,
    reason_code: str,
    explanation: str,
    expected_subcontractor: str,
    expected_quantity: Optional[float] = None,
    consumes: bool = False,
) -> AuditResult:
    if expected_quantity is None:
        submitted_code = match.final_po.canonical.get("submitted_item_code")
        expected_quantity = sum(
            item.canonical.get("expected_quantity", 0.0)
            for item in match.exact_item_records
            if item.canonical.get("expected_item_code") == submitted_code
        )
    return AuditResult(
        final_po=match.final_po,
        scope=match.scope,
        du_model_name=match.du_model_name,
        du_model_id=match.du_model_id,
        du_identity_key=match.du_identity_key,
        du_profile_ids=match.du_profile_ids,
        classification=classification,
        reason_code=reason_code,
        expected_items=match.expected_items,
        expected_subcontractor=expected_subcontractor,
        expected_quantity=expected_quantity,
        normal_quantity=0.0,
        duplicate_quantity=0.0,
        explanation=explanation,
        expected_evidence=match.evidence,
        matched_evidence=expected_evidence(match.exact_item_records),
        consumes_quantity=consumes,
    )


def site_keys(data: Dict[str, Any]) -> List[str]:
    return [key for key in [data.get("site_code"), data.get("du")] if key]


def entitlement_key(data: Dict[str, Any]) -> Tuple[str, str, str, str, str]:
    site_key = data.get("site_code") or data.get("du") or ""
    return (
        site_key,
        data.get("du_identity_key") or "UNKNOWN",
        data.get("scope") or "UNKNOWN",
        data.get("expected_item_code") or "",
        data.get("expected_subcontractor_norm") or "",
    )


def consumption_key(result: AuditResult) -> Tuple[str, str, str, str, str]:
    f = result.final_po.canonical
    site_key = f.get("site_code") or f.get("du") or ""
    return (
        site_key,
        result.du_identity_key or "UNKNOWN",
        result.scope,
        f.get("submitted_item_code") or "",
        normalize_subcontractor(result.expected_subcontractor),
    )


def expected_du_resolution_status(records: Sequence[ExpectedECCRecord]) -> str:
    statuses = {
        text(record.canonical.get("du_resolution_status")) or "UNKNOWN"
        for record in records
    }
    if "CONFLICT" in statuses:
        return "CONFLICT"
    if "UNKNOWN" in statuses:
        return "UNKNOWN"
    return "RESOLVED"


def expected_evidence(records: Sequence[ExpectedECCRecord]) -> str:
    if not records:
        return ""
    parts = []
    for record in records[:20]:
        parts.append(f"{Path(record.source_file).name}:{record.source_sheet}!{record.source_row}")
    if len(records) > 20:
        parts.append(f"... +{len(records) - 20} more")
    return "; ".join(parts)


def expected_item_text(records: Sequence[ExpectedECCRecord]) -> str:
    items = []
    for record in records:
        code = text(record.canonical.get("expected_item_code"))
        qty = record.canonical.get("expected_quantity", 0.0)
        if code:
            items.append(f"{code} x {qty:g}")
    return "; ".join(items)


def infer_scope_from_expected_or_final(records: Sequence[ExpectedECCRecord], final_po: FinalPORecord) -> str:
    scopes = [text(record.canonical.get("scope")) for record in records if text(record.canonical.get("scope"))]
    if scopes:
        return sorted(set(scopes))[0] if len(set(scopes)) == 1 else "MULTI"
    return infer_scope_from_final_po(final_po)


def infer_scope_from_ecc(source_file: str) -> str:
    name = Path(source_file).name.upper()
    if " TSS PR " in name:
        return "TSS"
    if " TI PR " in name:
        return "TI"
    if " PLANNING PR " in name:
        return "PLANNING"
    if " OPERATION PR " in name or " OPERATION BACKOFFICE PR " in name:
        return "OPERATION"
    return "UNKNOWN"


def infer_scope_from_final_po(record: FinalPORecord) -> str:
    data = record.canonical
    domain = text(data.get("business_domain")).lower()
    description = text(data.get("submitted_item_description")).lower()
    combined = f"{domain} {description}"
    if "planning" in combined:
        return "PLANNING"
    if "operation" in combined or "backoffice" in combined:
        return "OPERATION"
    if "survey" in combined or "tss" in combined:
        return "TSS"
    if "installation" in combined or "antenna" in combined or "microwave" in combined:
        return "TI"
    return "UNKNOWN"


def submitted_quantity(data: Dict[str, Any]) -> float:
    for field in ("submitted_quantity", "settlement_quantity", "paid_quantity"):
        qty = to_float(data.get(field), default=-1.0)
        if qty >= 0:
            return qty
    return 0.0


def source_headers(results: Sequence[AuditResult]) -> List[str]:
    if not results:
        return []
    return [header for header in results[0].final_po.raw.keys() if not header.startswith("_")]


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_code(value: Any) -> str:
    return text(value).upper()


def normalize_subcontractor(value: Any) -> str:
    raw = text(value).upper()
    raw = raw.replace("&", " AND ")
    raw = re.sub(
        r"\b(SDN|BHD|SDN\.|BHD\.|BERHAD|LTD|LIMITED|ENGINEERING|TECHNOLOGY|TECHNOLOGIES|SCIENCE|MALAYSIA)\b",
        " ",
        raw,
    )
    raw = re.sub(r"[^A-Z0-9]+", " ", raw).strip()
    if "GIROBUMI" in raw or raw == "GTSB":
        return "GTSB"
    if "GCI" in raw:
        return "GCI"
    if "ALL STAR" in raw or "ALLSTAR" in raw:
        return "ALLSTAR"
    return re.sub(r"\s+", " ", raw).strip()


def to_float(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return float(match.group(0)) if match else default


def dispatch_sort_key(data: Dict[str, Any]) -> Tuple[Any, str, str, int]:
    return (
        sortable_date(data.get("dispatch_date")),
        text(data.get("request_number")),
        text(data.get("dispatch_order_number")),
        int(to_float(data.get("po_line_number"), default=0.0)),
    )


def sortable_date(value: Any) -> Any:
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        return value
    value_text = text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value_text, fmt)
        except ValueError:
            pass
    return value_text


def column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def main() -> int:
    args = parse_args()
    try:
        summary = run_pipeline(args)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
