import importlib.util
import json
import sys
import tempfile
import unittest
from dataclasses import replace
from datetime import datetime
from pathlib import Path


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "scripts" / "audit_final_po.py"
SPEC = importlib.util.spec_from_file_location("audit_final_po", SCRIPT_PATH)
audit = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
sys.modules[SPEC.name] = audit
SPEC.loader.exec_module(audit)


def final_record(source_row=2, **overrides):
    canonical = {
        "dispatch_date": datetime(2026, 1, 2),
        "request_number": f"REQ-SYN-{source_row:03d}",
        "dispatch_order_number": f"DO-SYN-{source_row:03d}",
        "po_line_number": source_row,
        "site_code": "SITE-SYN-001",
        "du": "DU-SYN-001",
        "business_domain": "Planning",
        "submitted_item_code": "350001000403",
        "submitted_item_description": "Synthetic planning service",
        "submitted_quantity": 1.0,
        "settlement_quantity": 1.0,
        "submitted_subcontractor": "GCI",
    }
    canonical.update(overrides)
    canonical["site_code"] = audit.normalize_code(canonical.get("site_code"))
    canonical["du"] = audit.normalize_code(canonical.get("du"))
    canonical["submitted_item_code"] = audit.normalize_code(canonical.get("submitted_item_code"))
    canonical["submitted_subcontractor_norm"] = audit.normalize_subcontractor(canonical.get("submitted_subcontractor"))
    canonical["dispatch_sort_key"] = audit.dispatch_sort_key(canonical)
    return audit.FinalPORecord(source_row=source_row, raw=dict(canonical), canonical=canonical)


def expected_record(source_row=2, **overrides):
    source_file = overrides.pop("source_file", "Northern-GCI TX Mini Project Planning PR 20260706.xlsx")
    source_sheet = overrides.pop("source_sheet", "details")
    canonical = {
        "site_code": "SITE-SYN-001",
        "du": "DU-SYN-001",
        "region": "Northern",
        "expected_subcontractor": "GCI",
        "expected_item_code": "350001000403",
        "expected_item_description": "Synthetic planning service",
        "expected_quantity": 1.0,
        "scope": "PLANNING",
    }
    canonical.update(overrides)
    canonical["site_code"] = audit.normalize_code(canonical.get("site_code"))
    canonical["du"] = audit.normalize_code(canonical.get("du"))
    canonical["expected_item_code"] = audit.normalize_code(canonical.get("expected_item_code"))
    canonical["expected_quantity"] = audit.to_float(canonical.get("expected_quantity"))
    canonical["expected_subcontractor_norm"] = audit.normalize_subcontractor(canonical.get("expected_subcontractor"))
    canonical["entitlement_key"] = audit.entitlement_key(canonical)
    return audit.ExpectedECCRecord(
        source_file=source_file,
        source_sheet=source_sheet,
        source_row=source_row,
        raw=dict(canonical),
        canonical=canonical,
    )


def run_pipeline_for_records(final_records, expected_records):
    dataset = audit.CanonicalDataset(final_records, expected_records, metadata={})
    if any(
        "du_resolution_status" not in record.canonical
        for record in [*final_records, *expected_records]
    ):
        dataset = audit.canonical_builder(dataset)
    matches = audit.expected_matcher(dataset)
    audited = audit.audit_engine(matches, dataset.metadata)
    return audit.duplicate_resolver(audited).results


class TxPrAuditorTests(unittest.TestCase):
    def tearDown(self):
        audit.set_cancellation_probe(None)

    def test_large_pipeline_checks_cooperative_cancellation(self):
        audit.set_cancellation_probe(lambda: (_ for _ in ()).throw(RuntimeError("cancelled-by-test")))
        records = [final_record(source_row=index + 2, site_code=f"SITE-{index:04d}") for index in range(1000)]
        with self.assertRaisesRegex(RuntimeError, "cancelled-by-test"):
            audit.expected_matcher(audit.CanonicalDataset(records, [], {}))

    def test_du_registry_contains_nine_unique_create_pr_cd_identities(self):
        registry = audit.load_du_registry()
        identities = registry["identities"]

        self.assertEqual(len(identities), 9)
        self.assertEqual(len({item["identity_key"] for item in identities}), 9)
        self.assertEqual(
            {item["du_model_name"] for item in identities},
            {
                "TX Mini Project",
                "2023 TX Rollout",
                "MW EOS Swap",
                "2023 Celcomdigi BAU",
                "2024 Celcomdigi BAU",
                "Celcomdigi USP",
                "Jendela TX Migration",
                "ZTE TX MINI",
                "CD consolidation 2023",
            },
        )
        consolidation = next(
            item for item in identities if item["du_model_name"] == "CD consolidation 2023"
        )
        self.assertEqual(
            consolidation["profile_ids"],
            ["celcomdigi_cd_consolidation_2023_pr_v1"],
        )
        self.assertEqual(len(consolidation["view_ids"]), 2)
        tx_mini = next(item for item in identities if item["du_model_name"] == "TX Mini Project")
        self.assertEqual(tx_mini["profiles"][0]["profile_status"], "PRODUCTION")
        expected_contract_status = (
            "MATCHED"
            if Path(registry["source_contract_path"]).is_file()
            else "NOT_AVAILABLE"
        )
        self.assertEqual(registry["source_contract_status"], expected_contract_status)

    def test_du_registry_drift_from_create_pr_cd_fails_closed(self):
        registry = audit.load_du_registry()
        if not Path(registry["source_contract_path"]).is_file():
            self.skipTest("create-pr-cd source contract is not available")
        registry_data = json.loads(audit.DEFAULT_DU_REGISTRY.read_text(encoding="utf-8"))
        registry_data["identities"][0]["profiles"][0]["profile_status"] = "DRAFT"

        with tempfile.TemporaryDirectory() as tmpdir:
            registry_path = Path(tmpdir) / "du_registry.json"
            registry_path.write_text(json.dumps(registry_data), encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "differs from create-pr-cd"):
                audit.load_du_registry(registry_path)

    def test_all_nine_du_models_resolve_from_create_pr_cd_output_filenames(self):
        registry = audit.load_du_registry()

        for identity in registry["identities"]:
            with self.subTest(du_model=identity["du_model_name"]):
                source_file = (
                    f"Northern-GCI {identity['du_model_name']} "
                    "TSS PR 20260727.xlsx"
                )
                resolved = audit.resolve_du_identity(
                    registry,
                    source_text=source_file,
                )

                self.assertIsNotNone(resolved)
                self.assertEqual(resolved["identity_key"], identity["identity_key"])

    def test_canonical_builder_reports_nine_du_support_and_identifies_ecc_files(self):
        registry = audit.load_du_registry()
        expected = [
            expected_record(
                source_row=index + 2,
                source_file=(
                    f"Northern-GCI {identity['du_model_name']} "
                    "TSS PR 20260727.xlsx"
                ),
            )
            for index, identity in enumerate(registry["identities"])
        ]

        canonical = audit.canonical_builder(
            audit.CanonicalDataset([], expected, {}),
            registry,
        )

        self.assertEqual(
            {record.canonical["du_identity_key"] for record in canonical.expected_records},
            {item["identity_key"] for item in registry["identities"]},
        )
        self.assertEqual(canonical.metadata["du_support"]["supported_du_count"], 9)
        self.assertEqual(canonical.metadata["du_support"]["identified_ecc_file_count"], 9)
        self.assertEqual(canonical.metadata["du_support"]["unknown_ecc_file_count"], 0)

    def test_final_po_du_model_disambiguates_same_site_entitlement(self):
        registry = audit.load_du_registry()
        dataset = audit.canonical_builder(
            audit.CanonicalDataset(
                [final_record(project_name="MW EOS Swap")],
                [
                    expected_record(
                        source_file="Northern-GCI MW EOS Swap Planning PR 20260727.xlsx",
                        expected_quantity=1.0,
                    ),
                    expected_record(
                        source_file="Northern-GCI ZTE TX MINI Planning PR 20260727.xlsx",
                        expected_quantity=5.0,
                    ),
                ],
                {},
            ),
            registry,
        )

        results = run_pipeline_for_records(
            dataset.final_po_records,
            dataset.expected_records,
        )

        self.assertEqual(results[0].classification, "Normal")
        self.assertEqual(results[0].du_model_name, "MW EOS Swap")
        self.assertEqual(results[0].expected_quantity, 1.0)

    def test_final_po_resolves_du_from_logical_site_name_and_project_code(self):
        registry = audit.load_du_registry()
        dataset = audit.canonical_builder(
            audit.CanonicalDataset(
                [
                    final_record(
                        project_name="Malaysia CelcomDigi project",
                        project_code="P202202168750_D002",
                        logical_site_name="Wireless RAN/TX Mini Project/S00495_PORT",
                    )
                ],
                [],
                {},
            ),
            registry,
        )

        canonical = dataset.final_po_records[0].canonical
        self.assertEqual(canonical["du_model_name"], "TX Mini Project")
        self.assertEqual(canonical["du_project_key"], "Malaysia_CelcomDigi_Project")
        self.assertEqual(canonical["du_resolution_status"], "RESOLVED")

    def test_final_po_project_and_du_model_conflict_fails_closed(self):
        registry = audit.load_du_registry()
        dataset = audit.canonical_builder(
            audit.CanonicalDataset(
                [
                    final_record(
                        project_code="P202211283695_D002",
                        logical_site_name="Wireless RAN/TX Mini Project/S00495_PORT",
                    )
                ],
                [
                    expected_record(
                        source_file="Northern-GCI TX Mini Project Planning PR 20260727.xlsx"
                    )
                ],
                {},
            ),
            registry,
        )

        result = run_pipeline_for_records(
            dataset.final_po_records,
            dataset.expected_records,
        )[0]

        self.assertEqual(result.classification, "Abnormal - Invalid PO")
        self.assertEqual(result.reason_code, "INVALID_CONFLICTING_DU_IDENTITY")

    def test_unidentified_final_po_does_not_merge_multiple_du_entitlements(self):
        registry = audit.load_du_registry()
        dataset = audit.canonical_builder(
            audit.CanonicalDataset(
                [final_record(project_name="")],
                [
                    expected_record(
                        source_file="Northern-GCI MW EOS Swap Planning PR 20260727.xlsx",
                    ),
                    expected_record(
                        source_file="Northern-GCI ZTE TX MINI Planning PR 20260727.xlsx",
                    ),
                ],
                {},
            ),
            registry,
        )

        results = run_pipeline_for_records(
            dataset.final_po_records,
            dataset.expected_records,
        )

        self.assertEqual(results[0].classification, "Abnormal - Invalid PO")
        self.assertEqual(results[0].reason_code, "INVALID_AMBIGUOUS_DU_MODEL")
        self.assertEqual(results[0].du_identity_key, "MULTI")

    def test_unknown_ecc_du_identity_fails_closed(self):
        registry = audit.load_du_registry()
        dataset = audit.canonical_builder(
            audit.CanonicalDataset(
                [final_record(project_name="")],
                [
                    expected_record(
                        source_file="Northern-GCI Unregistered Model Planning PR 20260727.xlsx"
                    )
                ],
                {},
            ),
            registry,
        )

        results = run_pipeline_for_records(
            dataset.final_po_records,
            dataset.expected_records,
        )

        self.assertEqual(results[0].classification, "Abnormal - Invalid PO")
        self.assertEqual(results[0].reason_code, "INVALID_UNKNOWN_DU_MODEL")

    def test_conflicting_ecc_du_identity_evidence_fails_closed(self):
        registry = audit.load_du_registry()
        tx_mini = next(
            item for item in registry["identities"] if item["du_model_name"] == "TX Mini Project"
        )
        dataset = audit.canonical_builder(
            audit.CanonicalDataset(
                [final_record(project_name="")],
                [
                    expected_record(
                        source_file="Northern-GCI MW EOS Swap Planning PR 20260727.xlsx",
                        du_model_id=tx_mini["du_model_id"],
                    )
                ],
                {},
            ),
            registry,
        )

        results = run_pipeline_for_records(
            dataset.final_po_records,
            dataset.expected_records,
        )

        self.assertEqual(results[0].classification, "Abnormal - Invalid PO")
        self.assertEqual(results[0].reason_code, "INVALID_CONFLICTING_DU_IDENTITY")

    def test_cd_consolidation_preserves_exact_profile_and_view(self):
        registry = audit.load_du_registry()
        dataset = audit.canonical_builder(
            audit.CanonicalDataset(
                [],
                [
                    expected_record(
                        source_file=(
                            "Northern-GCI CD consolidation 2023 "
                            "Planning PR 20260727.xlsx"
                        ),
                        du_profile_id="celcomdigi_cd_consolidation_2023_pr_v1",
                        du_view_id="702960351133798763",
                    )
                ],
                {},
            ),
            registry,
        )

        canonical = dataset.expected_records[0].canonical
        self.assertEqual(
            canonical["du_profile_ids"],
            "celcomdigi_cd_consolidation_2023_pr_v1",
        )
        self.assertEqual(canonical["du_profile_statuses"], "DRAFT")
        self.assertEqual(canonical["du_view_ids"], "702960351133798763")

    def test_cd_consolidation_legacy_split_profile_fails_closed(self):
        registry = audit.load_du_registry()
        dataset = audit.canonical_builder(
            audit.CanonicalDataset(
                [final_record(project_name="CD consolidation 2023")],
                [
                    expected_record(
                        source_file=(
                            "Northern-GCI CD consolidation 2023 "
                            "Planning PR 20260727.xlsx"
                        ),
                        du_profile_id="cd_consolidation_2023_decom_pr_v1",
                        du_view_id="8359047522524230651",
                    )
                ],
                {},
            ),
            registry,
        )

        result = run_pipeline_for_records(
            dataset.final_po_records,
            dataset.expected_records,
        )[0]

        self.assertEqual(result.classification, "Abnormal - Invalid PO")
        self.assertEqual(result.reason_code, "INVALID_CONFLICTING_DU_IDENTITY")

    def test_quantity_consumption_is_isolated_per_du_model(self):
        registry = audit.load_du_registry()
        dataset = audit.canonical_builder(
            audit.CanonicalDataset(
                [
                    final_record(source_row=2, project_name="MW EOS Swap"),
                    final_record(source_row=3, project_name="ZTE TX MINI"),
                ],
                [
                    expected_record(
                        source_file="Northern-GCI MW EOS Swap Planning PR 20260727.xlsx",
                    ),
                    expected_record(
                        source_file="Northern-GCI ZTE TX MINI Planning PR 20260727.xlsx",
                    ),
                ],
                {},
            ),
            registry,
        )

        results = run_pipeline_for_records(
            dataset.final_po_records,
            dataset.expected_records,
        )

        self.assertEqual([result.classification for result in results], ["Normal", "Normal"])
        self.assertEqual(
            {result.du_model_name for result in results},
            {"MW EOS Swap", "ZTE TX MINI"},
        )

    def test_expected_quantity_is_isolated_by_scope(self):
        registry = audit.load_du_registry()
        dataset = audit.canonical_builder(
            audit.CanonicalDataset(
                [
                    final_record(
                        business_domain="Survey",
                        submitted_item_description="TSS survey",
                        submitted_quantity=2.0,
                    )
                ],
                [
                    expected_record(
                        source_file="Northern-GCI TX Mini Project TSS PR 20260727.xlsx",
                        expected_quantity=1.0,
                    ),
                    expected_record(
                        source_row=3,
                        source_file="Northern-GCI TX Mini Project TI PR 20260727.xlsx",
                        expected_quantity=5.0,
                    ),
                ],
                {},
            ),
            registry,
        )

        result = run_pipeline_for_records(
            dataset.final_po_records,
            dataset.expected_records,
        )[0]

        self.assertEqual(result.scope, "TSS")
        self.assertEqual(result.expected_quantity, 1.0)
        self.assertEqual(result.reason_code, "DUPLICATE_PARTIAL_QUANTITY")

    def test_unknown_final_scope_does_not_pool_multiple_scopes(self):
        registry = audit.load_du_registry()
        dataset = audit.canonical_builder(
            audit.CanonicalDataset(
                [
                    final_record(
                        business_domain="",
                        submitted_item_description="Unclassified service",
                    )
                ],
                [
                    expected_record(
                        source_file="Northern-GCI TX Mini Project TSS PR 20260727.xlsx",
                    ),
                    expected_record(
                        source_row=3,
                        source_file="Northern-GCI TX Mini Project TI PR 20260727.xlsx",
                    ),
                ],
                {},
            ),
            registry,
        )

        result = run_pipeline_for_records(
            dataset.final_po_records,
            dataset.expected_records,
        )[0]

        self.assertEqual(result.classification, "Abnormal - Invalid PO")
        self.assertEqual(result.reason_code, "INVALID_AMBIGUOUS_SCOPE")

    def test_expected_quantity_is_isolated_by_subcontractor(self):
        registry = audit.load_du_registry()
        dataset = audit.canonical_builder(
            audit.CanonicalDataset(
                [final_record(submitted_subcontractor="GCI", submitted_quantity=2.0)],
                [
                    expected_record(expected_subcontractor="GCI", expected_quantity=1.0),
                    expected_record(
                        source_row=3,
                        expected_subcontractor="OTHER",
                        expected_quantity=5.0,
                    ),
                ],
                {},
            ),
            registry,
        )

        result = run_pipeline_for_records(
            dataset.final_po_records,
            dataset.expected_records,
        )[0]

        self.assertEqual(result.expected_quantity, 1.0)
        self.assertEqual(result.expected_subcontractor, "GCI")
        self.assertEqual(result.reason_code, "DUPLICATE_PARTIAL_QUANTITY")

    def test_filter_final_po_period_keeps_only_matching_dispatch_month(self):
        january = final_record(source_row=2, dispatch_date=datetime(2026, 1, 3))
        february = final_record(source_row=3, dispatch_date="2026-02-04")
        invalid_date = final_record(source_row=4, dispatch_date="not-a-date")
        dataset = audit.CanonicalDataset([january, february, invalid_date], [], {})

        filtered = audit.filter_final_po_period(dataset, 2026, 2)

        self.assertEqual([record.source_row for record in filtered.final_po_records], [3])
        self.assertEqual(filtered.metadata["final_po_period_filter"]["input_row_count"], 3)
        self.assertEqual(filtered.metadata["final_po_period_filter"]["matched_row_count"], 1)

    def test_filter_final_po_period_requires_year_and_month_together(self):
        dataset = audit.CanonicalDataset([], [], {})
        with self.assertRaisesRegex(ValueError, "provided together"):
            audit.filter_final_po_period(dataset, 2026, None)

    def test_final_po_layout_auto_detects_supported_formats(self):
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl is required for workbook layout test")

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            for sheet_name, header_row in (("条目明细", 1), ("Sheet1", 2)):
                workbook_path = tmp_path / f"{sheet_name}.xlsx"
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name
                if header_row == 2:
                    ws.cell(1, 36, "PM/TL to feedback")
                ws.cell(header_row, 1, "派工日期")
                ws.cell(header_row, 2, "派工单号")
                ws.cell(header_row + 1, 1, datetime(2026, 1, 2))
                ws.cell(header_row + 1, 2, "DO-SYN-001")
                wb.save(workbook_path)

                resolved = audit.resolve_final_po_layout(workbook_path, None, None)
                rows, metadata = audit.read_table(workbook_path, *resolved)

                self.assertEqual(resolved, (sheet_name, header_row))
                self.assertEqual(metadata["sheet"], sheet_name)
                self.assertEqual(metadata["header_row"], header_row)
                self.assertEqual(rows[0]["派工单号"], "DO-SYN-001")

    def test_final_po_layout_preserves_explicit_overrides(self):
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl is required for workbook layout test")

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "custom.xlsx"
            wb = Workbook()
            wb.active.title = "Custom"
            wb.save(workbook_path)

            self.assertEqual(
                audit.resolve_final_po_layout(workbook_path, "Custom", 3),
                ("Custom", 3),
            )

    def test_valid_row_is_normal_against_create_pr_cd_output(self):
        results = run_pipeline_for_records([final_record()], [expected_record()])

        self.assertEqual(results[0].classification, "Normal")
        self.assertEqual(results[0].reason_code, "NORMAL_FULL")
        self.assertEqual(results[0].normal_quantity, 1.0)

    def test_wrong_item_when_site_exists_but_pbom_not_expected(self):
        results = run_pipeline_for_records(
            [final_record(submitted_item_code="WRONG-CODE")],
            [expected_record()],
        )

        self.assertEqual(results[0].classification, "Abnormal - Wrong PO")
        self.assertEqual(results[0].reason_code, "WRONG_LINE_ITEM_MAPPING")

    def test_missing_generated_ecc_entitlement_is_invalid(self):
        results = run_pipeline_for_records(
            [final_record(site_code="SITE-NOT-GENERATED", du="DU-NOT-GENERATED")],
            [expected_record()],
        )

        self.assertEqual(results[0].classification, "Abnormal - Invalid PO")
        self.assertEqual(results[0].reason_code, "INVALID_NOT_IN_CREATE_PR_CD_OUTPUT")

    def test_subcontractor_changed_is_invalid(self):
        results = run_pipeline_for_records(
            [final_record(submitted_subcontractor="Other Supplier")],
            [expected_record()],
        )

        self.assertEqual(results[0].classification, "Abnormal - Invalid PO")
        self.assertEqual(results[0].reason_code, "INVALID_SUBCON_CHANGED")

    def test_duplicate_consumption_uses_dispatch_order(self):
        first = final_record(source_row=2, dispatch_date=datetime(2026, 1, 1), request_number="REQ-001")
        second = final_record(source_row=3, dispatch_date=datetime(2026, 1, 2), request_number="REQ-002")

        results = run_pipeline_for_records([second, first], [expected_record(expected_quantity=1.0)])

        self.assertEqual(results[0].classification, "Abnormal - Duplicate PO")
        self.assertEqual(results[0].reason_code, "DUPLICATE_FULL_QUANTITY")
        self.assertEqual(results[1].classification, "Normal")

    def test_partial_duplicate_preserves_normal_and_duplicate_quantities(self):
        submitted = final_record(submitted_quantity=2.0, settlement_quantity=2.0)
        results = run_pipeline_for_records([submitted], [expected_record(expected_quantity=1.5)])

        self.assertEqual(results[0].classification, "Abnormal - Duplicate PO")
        self.assertEqual(results[0].reason_code, "DUPLICATE_PARTIAL_QUANTITY")
        self.assertEqual(results[0].normal_quantity, 1.5)
        self.assertEqual(results[0].duplicate_quantity, 0.5)

    def test_annotation_status_aggregation(self):
        normal = run_pipeline_for_records([final_record()], [expected_record()])[0]
        duplicate = run_pipeline_for_records(
            [
                final_record(source_row=2, request_number="REQ-001"),
                final_record(source_row=3, request_number="REQ-002"),
            ],
            [expected_record(expected_quantity=1.0)],
        )[1]

        self.assertEqual(audit.annotation_values([normal])[0], "NORMAL")
        self.assertEqual(audit.annotation_values([])[0], "NOT_IN_FINAL_PO")
        self.assertEqual(audit.annotation_values([duplicate])[0], "DUPLICATE")
        self.assertEqual(audit.annotation_values([normal, duplicate])[0], "MIXED")

    def test_annotated_ecc_writer_copies_without_touching_source(self):
        try:
            from openpyxl import Workbook, load_workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl is required for workbook annotation test")

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            source = tmp_path / "Northern-GCI TX Mini Project Planning PR 20260706.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "details"
            headers = [
                "SN.",
                "Purchasing Area*",
                "Region*",
                "Site ID*",
                "Site Name*",
                "Delivery Unit Code*",
                "Logical Site Name",
                "Contract Number *",
                "Subcontractor*",
                "PBOM Code*",
                "SOW*",
                "Unit*",
                "Quantity*",
                "Remarks",
                None,
                "Contract Number",
            ]
            for idx, header in enumerate(headers, 1):
                ws.cell(1, idx, header)
            ws.cell(2, 1, 1)
            ws.cell(2, 4, "SITE-SYN-001")
            ws.cell(2, 6, "DU-SYN-001")
            ws.cell(2, 9, "GCI")
            ws.cell(2, 10, "350001000403")
            ws.cell(2, 13, 1)
            wb.save(source)

            result = run_pipeline_for_records(
                [final_record()],
                [expected_record(source_file=str(source), source_row=2)],
            )[0]
            summary = audit.annotated_ecc_writer(
                audit.AuditDataset([result], {}),
                [source],
                tmp_path / "output",
                "TEST_RUN",
                "details",
            )

            source_wb = load_workbook(source, read_only=True, data_only=True)
            source_headers = [cell.value for cell in next(source_wb["details"].iter_rows(min_row=1, max_row=1))]
            self.assertNotIn("Audit Status", source_headers)
            source_wb.close()

            copied = tmp_path / "output" / "TEST_RUN" / source.name
            copied_wb = load_workbook(copied, read_only=True, data_only=True)
            copied_ws = copied_wb["details"]
            copied_headers = [cell.value for cell in next(copied_ws.iter_rows(min_row=1, max_row=1))]
            self.assertIn("Audit Status", copied_headers)
            self.assertEqual(copied_ws.cell(2, 17).value, "NORMAL")
            self.assertEqual(Path(summary["files"][0]["output_file"]).name, source.name)
            copied_wb.close()


if __name__ == "__main__":
    unittest.main()
