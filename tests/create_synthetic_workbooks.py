#!/usr/bin/env python
import argparse
import importlib.util
import sys
from datetime import datetime
from pathlib import Path


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "scripts" / "audit_final_po.py"
SPEC = importlib.util.spec_from_file_location("audit_final_po", SCRIPT_PATH)
audit = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
sys.modules[SPEC.name] = audit
SPEC.loader.exec_module(audit)


def parse_args():
    parser = argparse.ArgumentParser(description="Create synthetic TX PR Auditor smoke-test workbooks.")
    parser.add_argument("--output-dir", required=True)
    return parser.parse_args()


def write_workbooks(output_dir: Path):
    from openpyxl import Workbook

    output_dir.mkdir(parents=True, exist_ok=True)

    final_po = Workbook()
    ws = final_po.active
    ws.title = audit.FINAL_PO_SHEET_NAME
    final_headers = list(audit.FINAL_PO_FIELD_MAP.keys())
    ws.append(final_headers)
    final_values = {field: "" for field in audit.FINAL_PO_FIELD_MAP.values()}
    final_values.update(
        {
            "dispatch_date": datetime(2026, 1, 2),
            "dispatch_order_number": "DO-SYN-001",
            "po_line_number": 1,
            "request_number": "REQ-SYN-001",
            "project_name": "Synthetic Fiber Trial",
            "project_code": "PRJ-SYN-001",
            "business_domain": "Planning",
            "region": "Synthetic Region",
            "purchasing_area": "Synthetic Area",
            "submitted_subcontractor": "Synthetic Vendor Alpha",
            "logical_site_name": "Synthetic Hill",
            "du": "DU-SYN-001",
            "physical_site_name": "Synthetic Hill Physical",
            "site_code": "SITE-SYN-001",
            "submitted_item_code": audit.PLANNING_ITEM_CODE,
            "submitted_item_description": "Synthetic planning service",
            "submitted_unit": "Each",
            "submitted_quantity": 1,
            "settlement_quantity": 1,
            "paid_quantity": 0,
            "product_model_remark": "Synthetic model",
            "dispatch_status": "Synthetic approved",
            "subcontractor_code": "SUB-SYN-001",
        }
    )
    ws.append([final_values[audit.FINAL_PO_FIELD_MAP[header]] for header in final_headers])
    final_po.save(output_dir / "Final_PO.synthetic.xlsx")

    epms = Workbook()
    ws = epms.active
    ws.title = audit.EPMS_SHEET_NAME
    ws.append([])
    ws.append([])
    ws.append([])
    epms_headers = list(audit.EPMS_FIELD_MAP.keys())
    ws.append(epms_headers)
    epms_values = {field: "" for field in audit.EPMS_FIELD_MAP.values()}
    epms_values.update(
        {
            "site_code": "SITE-SYN-001",
            "site_name": "Synthetic Hill",
            "du": "DU-SYN-001",
            "epms_region": "Synthetic Region",
            "province_state": "Synthetic State",
            "latitude": "",
            "longitude": "",
            "tx_upgrade_scope": "Synthetic Upgrade",
            "boq_configuration": "Synthetic BOQ",
            "tx_sow": "Synthetic Planning",
            "tx_sow_details": "Synthetic details",
            "ne_sow_details": "Synthetic NE details",
            "fe_sow_details": "Synthetic FE details",
            "antenna_size_ne": "",
            "antenna_size_fe": "",
            "expected_tss_subcontractor": "Synthetic Vendor Alpha",
            "existing_tss_pr": "",
            "expected_ti_subcontractor": "Synthetic Vendor Beta",
            "existing_ti_pr": "",
            "tx_cutover_date": "",
            "expected_planning_subcontractor": "Alpha Field Services",
            "existing_planning_pr": "",
        }
    )
    ws.append([epms_values[audit.EPMS_FIELD_MAP[header]] for header in epms_headers])
    epms.save(output_dir / "EPMS.synthetic.xlsx")

    pr_model = Workbook()
    ws = pr_model.active
    ws.title = "TX Line Item (After 21-Apr 26)"
    ws.append(["TSS Model"])
    ws.append(["Synthetic TSS", "TSS-SYN-001", "Synthetic TSS service", "Each", 1, "Mandatory"])
    ws.append(["TI Model"])
    ws.append(["Synthetic TI", "TI-SYN-001", "Synthetic TI service", "Each", 1, "Mandatory"])
    pr_model.save(output_dir / "pr_model.synthetic.xlsx")


def main():
    write_workbooks(Path(parse_args().output_dir))


if __name__ == "__main__":
    main()
